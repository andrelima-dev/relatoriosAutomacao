import os
from datetime import date, datetime
from typing import Callable

import pandas as pd
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

from core.utils import is_date_col, get_col

ABA_NOME = "Planilha1"

CATEGORIAS_ATIVAS = {"ADVOGADO", "SUPLEMENTAR"}
SITUACOES_ATIVAS = {
    "ATIVO COM IMPEDIMENTO",
    "ATIVO PLENO",
    "EXECUTADO",
    "SUSPENSO EM OUTRA SECCIONAL",
    "SUSPENSO POR PROCESSO",
}

_thin = Side(style="thin")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_font_header = Font(name="Verdana", size=8, bold=True)
_font_data = Font(name="Verdana", size=8)

_DATE_FORMATS = (
    "%Y-%m-%dT%H:%M:%S.%fZ",
    "%Y-%m-%dT%H:%M:%SZ",
    "%Y-%m-%dT%H:%M:%S",
    "%d/%m/%Y",
    "%Y-%m-%d",
    "%d-%m-%Y",
    "%Y/%m/%d",
)


def _parse_date(value: str) -> datetime | None:
    if not value or not isinstance(value, str):
        return None
    v = value.strip()
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(v, fmt)
        except ValueError:
            continue
    return None


def _aplicar_filtro_data(
    df: pd.DataFrame,
    data_col: str | None,
    data_inicio: datetime | None,
    data_fim: datetime | None,
) -> pd.DataFrame:
    if not data_col or (data_inicio is None and data_fim is None):
        return df
    col = get_col(df, data_col)
    if col is None:
        return df
    parsed = df[col].map(_parse_date)
    mask = pd.Series([True] * len(df), index=df.index)
    if data_inicio:
        mask &= parsed >= data_inicio
    if data_fim:
        mask &= parsed <= data_fim
    return df[mask].reset_index(drop=True)


def _escrever_planilha(
    ws,
    df: pd.DataFrame,
    progress_cb: Callable[[int], None] | None = None,
    prog_start: int = 0,
    prog_end: int = 100,
):
    date_col_indices = {i for i, c in enumerate(df.columns) if is_date_col(c)}
    n_rows = len(df)

    # Pre-computar larguras antes de escrever (necessário no write_only mode)
    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        max_header = len(str(col_name))
        max_data = (
            df.iloc[:, col_idx - 1].astype(str).str.len().max()
            if n_rows > 0 else 0
        )
        ws.column_dimensions[col_letter].width = min(
            max(max_header, int(max_data), 0) + 2, 50
        )

    # Cabeçalho
    header = []
    for col_name in df.columns:
        c = WriteOnlyCell(ws, value=col_name)
        c.font = _font_header
        c.border = _border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        header.append(c)
    ws.append(header)

    # Dados — df.to_numpy() evita overhead do itertuples por linha
    arr = df.to_numpy()
    for row_idx, row in enumerate(arr):
        cells = []
        for col_idx, value in enumerate(row):
            raw = "" if (value is None or (isinstance(value, float) and value != value)) else value
            if col_idx in date_col_indices:
                parsed = _parse_date(str(raw))
                c = WriteOnlyCell(ws, value=parsed if parsed else (raw or None))
                if parsed:
                    c.number_format = "DD/MM/YYYY"
            else:
                c = WriteOnlyCell(ws, value=raw if raw != "" else None)
            c.font = _font_data
            c.border = _border
            c.alignment = Alignment(vertical="center")
            cells.append(c)
        ws.append(cells)

        if progress_cb and n_rows > 0 and row_idx % 500 == 0:
            pct = prog_start + int((row_idx / n_rows) * (prog_end - prog_start))
            progress_cb(pct)

    if progress_cb:
        progress_cb(prog_end)


def _nome_arquivo(nome_base: str, sufixo: str) -> str:
    hoje = date.today().strftime("%d-%m-%Y")
    return f"{nome_base}{sufixo} {hoje}.xlsx"


def _criar_wb() -> tuple:
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title=ABA_NOME)
    return wb, ws


def _salvar_wb(wb, path: str, log_cb: Callable[[str, str], None]) -> str:
    """Salva o workbook. Se o arquivo estiver aberto/bloqueado (ex.: aberto no
    Excel), salva com um nome alternativo em vez de quebrar a geração."""
    try:
        wb.save(path)
        return path
    except PermissionError:
        base, ext = os.path.splitext(path)
        for i in range(2, 100):
            alt = f"{base} ({i}){ext}"
            try:
                wb.save(alt)
                log_cb(
                    f"'{os.path.basename(path)}' estava aberto no Excel — "
                    f"salvo como '{os.path.basename(alt)}'. Feche o arquivo para "
                    f"sobrescrever na próxima vez.",
                    "info",
                )
                return alt
            except PermissionError:
                continue
        raise PermissionError(
            f"Não foi possível salvar '{os.path.basename(path)}'. "
            f"Feche o arquivo no Excel e gere novamente."
        )


def gerar_relatorios(
    df: pd.DataFrame,
    pasta_saida: str,
    log_cb: Callable[[str, str], None],
    gerar_geral: bool = True,
    gerar_ativos: bool = True,
    gerar_por_uf: bool = False,
    gerar_por_comarca: bool = False,
    categorias_filtro: set[str] | None = None,
    situacoes_filtro: set[str] | None = None,
    uf_col: str | None = None,
    uf_filtro: list[str] | None = None,
    comarca_col: str | None = None,
    comarca_filtro: list[str] | None = None,
    data_col: str | None = None,
    data_inicio: datetime | None = None,
    data_fim: datetime | None = None,
    nome_base: str = "RELATORIO CADASTRO ADVOGADOS GERAL",
    progress_cb: Callable[[int], None] | None = None,
    gerar_personalizado: bool = False,
    colunas_personalizado: list[str] | None = None,
    uf_personalizado: list[str] | None = None,
    filtros_valores_personalizado: dict[str, list[str]] | None = None,
    personalizado_base: str = "ativos",
) -> dict[str, str | None]:
    os.makedirs(pasta_saida, exist_ok=True)

    cats = {c.upper() for c in (categorias_filtro if categorias_filtro is not None else CATEGORIAS_ATIVAS)}
    sits = {s.upper() for s in (situacoes_filtro if situacoes_filtro is not None else SITUACOES_ATIVAS)}

    df_base = _aplicar_filtro_data(df, data_col, data_inicio, data_fim)

    tem_personalizado = gerar_personalizado and bool(colunas_personalizado)
    n_reports = sum([
        gerar_geral,
        gerar_ativos,
        gerar_por_uf and bool(uf_filtro),
        gerar_por_comarca and bool(comarca_filtro),
        tem_personalizado,
    ])
    chunk = 100 // max(n_reports, 1)
    prog_offset = 0

    results: dict[str, str | None] = {
        "geral": None, "ativos": None, "por_uf": None, "por_comarca": None, "personalizado": None,
    }

    if gerar_geral:
        log_cb("Gerando GERAL...", "info")
        path = os.path.join(pasta_saida, _nome_arquivo(nome_base, ""))
        wb, ws = _criar_wb()
        _escrever_planilha(ws, df_base, progress_cb, prog_offset, prog_offset + chunk)
        path = _salvar_wb(wb, path, log_cb)
        log_cb(f"Gerando GERAL... ✔ ({len(df_base)} registros)", "ok")
        results["geral"] = path
        prog_offset += chunk

    if gerar_ativos:
        log_cb("Gerando GERAL ATIVOS...", "info")
        col_cat = get_col(df_base, "CATEGORIA")
        col_sit = get_col(df_base, "SITUACAO_INSCRICAO")
        if col_cat is None:
            raise ValueError("Coluna 'CATEGORIA' não encontrada no XML.")
        if col_sit is None:
            raise ValueError("Coluna 'SITUACAO_INSCRICAO' não encontrada no XML.")

        mask = (
            df_base[col_cat].str.strip().str.upper().isin(cats)
            & df_base[col_sit].str.strip().str.upper().isin(sits)
        )
        df_ativos = df_base[mask].reset_index(drop=True)

        path = os.path.join(pasta_saida, _nome_arquivo(nome_base, " ATIVOS"))
        wb, ws = _criar_wb()
        _escrever_planilha(ws, df_ativos, progress_cb, prog_offset, prog_offset + chunk)
        path = _salvar_wb(wb, path, log_cb)
        log_cb(f"Gerando GERAL ATIVOS... ✔ ({len(df_ativos)} registros)", "ok")
        results["ativos"] = path
        prog_offset += chunk

    if gerar_por_uf and uf_filtro:
        col_uf_real = get_col(df_base, uf_col or "UF")
        if col_uf_real is None:
            raise ValueError(f"Coluna UF/Seccional '{uf_col or 'UF'}' não encontrada no XML.")

        col_cat = get_col(df_base, "CATEGORIA")
        col_sit = get_col(df_base, "SITUACAO_INSCRICAO")
        if col_cat and col_sit:
            mask_ativos = (
                df_base[col_cat].str.strip().str.upper().isin(cats)
                & df_base[col_sit].str.strip().str.upper().isin(sits)
            )
            df_filtrado = df_base[mask_ativos].reset_index(drop=True)
        else:
            df_filtrado = df_base

        chunk_uf = chunk // max(len(uf_filtro), 1)
        paths_uf = []
        for label in uf_filtro:
            label = label.strip()
            log_cb(f"Gerando subseção {label}...", "info")
            mask = df_filtrado[col_uf_real].str.strip().str.upper() == label.upper()
            df_uf = df_filtrado[mask].reset_index(drop=True)
            sufixo_arquivo = label.upper().replace("/", "-").replace("\\", "-")
            path = os.path.join(pasta_saida, _nome_arquivo(nome_base, f" {sufixo_arquivo}"))
            wb, ws = _criar_wb()
            _escrever_planilha(ws, df_uf, progress_cb, prog_offset, prog_offset + chunk_uf)
            path = _salvar_wb(wb, path, log_cb)
            log_cb(f"Gerando subseção {label}... ✔ ({len(df_uf)} registros)", "ok")
            paths_uf.append(path)
            prog_offset += chunk_uf
        results["por_uf"] = paths_uf or None

    if gerar_por_comarca and comarca_filtro:
        col_comarca_real = get_col(df_base, comarca_col or "COMARCA")
        if col_comarca_real is None:
            raise ValueError("Coluna de município não encontrada no XML.")

        col_cat = get_col(df_base, "CATEGORIA")
        col_sit = get_col(df_base, "SITUACAO_INSCRICAO")
        if col_cat and col_sit:
            mask_ativos = (
                df_base[col_cat].str.strip().str.upper().isin(cats)
                & df_base[col_sit].str.strip().str.upper().isin(sits)
            )
            df_filtrado_comarca = df_base[mask_ativos].reset_index(drop=True)
        else:
            df_filtrado_comarca = df_base

        municipios_upper = {m.strip().upper() for m in comarca_filtro}
        mask = df_filtrado_comarca[col_comarca_real].str.strip().str.upper().isin(municipios_upper)
        df_municipios = df_filtrado_comarca[mask].reset_index(drop=True)

        n = len(comarca_filtro)
        if n == 1:
            sufixo_arq = comarca_filtro[0].strip().upper().replace("/", "-").replace("\\", "-")
            sufixo = f" MUNICIPIO {sufixo_arq}"
        else:
            sufixo = f" MUNICIPIOS ({n})"

        log_cb(f"Gerando por município ({n} selecionado(s))...", "info")
        path = os.path.join(pasta_saida, _nome_arquivo(nome_base, sufixo))
        wb, ws = _criar_wb()
        _escrever_planilha(ws, df_municipios, progress_cb, prog_offset, prog_offset + chunk)
        path = _salvar_wb(wb, path, log_cb)
        log_cb(f"Gerando por município... ✔ ({len(df_municipios)} registros)", "ok")
        results["por_comarca"] = path
        prog_offset += chunk

    if tem_personalizado:
        cols_validas = [c for c in colunas_personalizado if c in df_base.columns]
        if not cols_validas:
            log_cb("Personalizado: nenhuma coluna válida selecionada.", "erro")
        else:
            # Base do personalizado: "ativos" (padrão) filtra ativos; "geral" usa tudo
            col_cat = get_col(df_base, "CATEGORIA")
            col_sit = get_col(df_base, "SITUACAO_INSCRICAO")
            if personalizado_base == "geral":
                df_ativos_base = df_base
                log_cb("Personalizado: base GERAL (todos os registros)", "info")
            elif col_cat and col_sit:
                mask_ativos = (
                    df_base[col_cat].str.strip().str.upper().isin(cats)
                    & df_base[col_sit].str.strip().str.upper().isin(sits)
                )
                df_ativos_base = df_base[mask_ativos].reset_index(drop=True)
                log_cb(f"Personalizado: base ATIVOS ({len(df_ativos_base)} registros)", "info")
            else:
                df_ativos_base = df_base

            # Filtros de valores por coluna (estilo AutoFiltro do Excel)
            if filtros_valores_personalizado:
                for col_nome, valores in filtros_valores_personalizado.items():
                    if not valores:
                        continue
                    col_real = get_col(df_ativos_base, col_nome)
                    if col_real is None or col_real not in df_ativos_base.columns:
                        continue
                    vals_upper = {str(v).strip().upper() for v in valores}
                    mask_val = df_ativos_base[col_real].astype(str).str.strip().str.upper().isin(vals_upper)
                    df_ativos_base = df_ativos_base[mask_val].reset_index(drop=True)
                    log_cb(
                        f"Personalizado: filtro '{col_nome}' "
                        f"({len(valores)} valor(es)) → {len(df_ativos_base)} registros",
                        "info",
                    )

            df_custom = df_ativos_base[cols_validas].copy()
            log_cb("Gerando PERSONALIZADO...", "info")
            path = os.path.join(pasta_saida, _nome_arquivo(nome_base, " PERSONALIZADO"))
            wb, ws = _criar_wb()
            _escrever_planilha(ws, df_custom, progress_cb, prog_offset, prog_offset + chunk)
            path = _salvar_wb(wb, path, log_cb)
            log_cb(
                f"Gerando PERSONALIZADO... ✔ ({len(df_custom)} registros, {len(cols_validas)} colunas)",
                "ok",
            )
            results["personalizado"] = path
            prog_offset += chunk

    if progress_cb:
        progress_cb(100)

    return results
